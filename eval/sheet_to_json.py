# -*- coding: utf-8 -*-
"""
sheet_to_json.py — prevedie vyplneny XLSX harok na JSON pre run_eval.py.

Spustenie:
    python3 sheet_to_json.py [vstup.xlsx] [vystup.json]
Predvolene: D9_zlata_sada.xlsx -> D9_zlata_sada.json

Do vystupu idu IBA riadky so stavom "hotovo". Rozpracovane a vyradene
sa preskocia a vypise sa, kolko ich bolo — sada sa da merat priebezne.
"""
import json, os, sys
from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = sys.argv[1] if len(sys.argv) > 1 else os.path.join(HERE, "D9_zlata_sada.xlsx")
DST = sys.argv[2] if len(sys.argv) > 2 else os.path.join(HERE, "D9_zlata_sada.json")

COL = {  # nazov -> index stlpca (1-based), podla build_sheet.COLS
    "id":1, "question":2, "searchMode":3, "sectionKey":4, "companyCode":5,
    "accessLevel":6, "precedenceRule":7, "trapType":8, "expectedBehaviour":9,
    "goldAnswer":10, "src1doc":11, "src1ref":12, "src2doc":13, "src2ref":14,
    "status":15, "notes":16,
}

def val(ws, r, key):
    v = ws.cell(r, COL[key]).value
    return v.strip() if isinstance(v, str) else v

def main():
    wb = load_workbook(SRC, data_only=True)
    ws = wb["Sada"]

    out, skipped, problems = [], {"prazdne":0, "rozpracovane":0, "vyradene":0}, []

    for r in range(3, ws.max_row + 1):
        qid = val(ws, r, "id")
        if not qid or qid == "PRÍKLAD":
            continue
        status = (val(ws, r, "status") or "").lower()
        if status == "vyradene":
            skipped["vyradene"] += 1; continue
        if status == "rozpracovane":
            skipped["rozpracovane"] += 1; continue
        if status != "hotovo":
            skipped["prazdne"] += 1; continue

        gold = val(ws, r, "goldAnswer")
        behaviour = val(ws, r, "expectedBehaviour")
        sources = []
        for d, s in (("src1doc","src1ref"), ("src2doc","src2ref")):
            doc, ref = val(ws, r, d), val(ws, r, s)
            if doc:
                sources.append({"document": doc, "articleRef": ref or None})

        # kontrola konzistencie — nech sa chyby najdu tu, nie az pri merani
        if not gold:
            problems.append(f"{qid}: stav 'hotovo', ale chyba odpoved")
        if behaviour == "answer" and not sources:
            problems.append(f"{qid}: vecna odpoved bez uvedeneho zdroja")
        if val(ws, r, "precedenceRule") and behaviour == "answer" and len(sources) < 2:
            problems.append(f"{qid}: otazka na precedenciu ma mat OBA predpisy")

        out.append({
            "id": qid,
            "question": val(ws, r, "question"),
            "searchMode": val(ws, r, "searchMode"),
            "sectionKey": None if val(ws, r, "sectionKey") in (None, "—") else val(ws, r, "sectionKey"),
            "companyCode": val(ws, r, "companyCode"),
            "accessLevel": val(ws, r, "accessLevel"),
            "precedenceRule": val(ws, r, "precedenceRule") or None,
            "trapType": val(ws, r, "trapType") or None,
            "expectedBehaviour": behaviour,
            "goldAnswer": gold,
            "goldSources": sources,
            "notes": val(ws, r, "notes") or "",
        })

    json.dump(out, open(DST, "w", encoding="utf-8"), ensure_ascii=False, indent=2)

    print(f"hotovych otazok:  {len(out)}  ->  {DST}")
    print(f"preskocene:       rozpracovane {skipped['rozpracovane']} · "
          f"nevyplnene {skipped['prazdne']} · vyradene {skipped['vyradene']}")
    if problems:
        print(f"\nUPOZORNENIA ({len(problems)}):")
        for p in problems: print("  -", p)
    if len(out) < 50:
        print(f"\nPOZOR: rozhodnutie D9 ziada 50-100 otazok, hotovych je {len(out)}.")

if __name__ == "__main__":
    main()
