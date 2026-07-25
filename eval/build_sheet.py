# -*- coding: utf-8 -*-
"""
build_sheet.py — vygeneruje XLSX harok zlatej sady D9 pre legislativca.

Vstup:  seed/questions_seed.json
Vystup: D9_zlata_sada.xlsx

Spustenie:  python3 build_sheet.py
"""
import json, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
SEED = os.path.join(HERE, "seed", "questions_seed.json")
OUT  = os.path.join(HERE, "D9_zlata_sada.xlsx")

FONT = "Arial"
INK    = Font(name=FONT, size=10)
BOLD   = Font(name=FONT, size=10, bold=True)
HEADF  = Font(name=FONT, size=10, bold=True, color="FFFFFF")
TITLE  = Font(name=FONT, size=14, bold=True)
MUTED  = Font(name=FONT, size=9, color="666666")

HEAD_FILL   = PatternFill("solid", fgColor="11151C")   # tmava hlavicka
LOCK_FILL   = PatternFill("solid", fgColor="ECEEF1")   # predvyplnene, needitovat
INPUT_FILL  = PatternFill("solid", fgColor="FFF7CC")   # zlte = vyplnit
TRAP_FILL   = PatternFill("solid", fgColor="FCE8E6")   # pasca
THIN = Side(style="thin", color="DDE1E7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP = Alignment(vertical="top", wrap_text=True)

COLS = [
    ("id",                "ID",                        9,  "lock"),
    ("question",          "Otázka",                   46,  "semi"),
    ("searchMode",        "Typ vyhľadávania",         15,  "lock"),
    ("sectionKey",        "Sekcia",                   19,  "lock"),
    ("companyCode",       "Organizácia",              12,  "lock"),
    ("accessLevel",       "Prístup",                  10,  "lock"),
    ("precedenceRule",    "Precedencia",              12,  "lock"),
    ("trapType",          "Typ pasce",                18,  "lock"),
    ("expectedBehaviour", "Očakávané správanie",      18,  "lock"),
    ("goldAnswer",        "OVERENÁ ODPOVEĎ",          52,  "input"),
    ("src1doc",           "Zdroj 1 — dokument",       26,  "input"),
    ("src1ref",           "Zdroj 1 — § / článok",     18,  "input"),
    ("src2doc",           "Zdroj 2 — dokument",       26,  "input"),
    ("src2ref",           "Zdroj 2 — § / článok",     18,  "input"),
    ("status",            "Stav",                     13,  "input"),
    ("notes",             "Poznámka legislatívca",    34,  "input"),
]

def build():
    rows = json.load(open(SEED, encoding="utf-8"))
    wb = Workbook()

    # ══════════ 1. LEGENDA ══════════
    lg = wb.active
    lg.title = "Legenda"
    lg.sheet_view.showGridLines = False
    lg["A1"] = "Contineo — zlatá sada D9"; lg["A1"].font = TITLE
    lines = [
        ("", ""),
        ("Čo je toto", "Sada overených otázok a odpovedí, proti ktorej sa meria kvalita vyhľadávania pred spustením."),
        ("", "Bez nej je porovnanie modelov dohad. Odpovede musia byť overené človekom, nie vygenerované."),
        ("", ""),
        ("Čo treba vyplniť", "Iba ŽLTÉ stĺpce (J–P) na hárku „Sada“. Sivé stĺpce sú predvyplnené — needitovať."),
        ("", "Otázku (stĺpec B) môžete upraviť alebo preformulovať, ak je nepresná alebo neprirodzená."),
        ("", ""),
        ("OVERENÁ ODPOVEĎ", "Krátka vecná odpoveď, ako by ju mal dať systém. 2–5 viet. Bez úvodných fráz."),
        ("Zdroj 1 / Zdroj 2", "Presný názov predpisu a § alebo článok, kde odpoveď stojí. Ak stačí jeden zdroj, druhý nechajte prázdny."),
        ("", "Pri otázkach na precedenciu (R1–R4) uveďte OBA predpisy — všeobecný aj špecifický."),
        ("Stav", "hotovo = odpoveď overená · rozpracovane = treba doriešiť · vyradene = otázka nedáva zmysel"),
        ("Poznámka", "Čokoľvek, čo má vedieť ten, kto bude vyhodnocovať. Napr. „sporné, treba potvrdiť komisiou“."),
        ("", ""),
        ("Pozor — pasce", "Riadky s červeným „Typ pasce“ sú zámerné. Systém na ne NEMÁ odpovedať vecne."),
        ("  out_of_domain", "Mimo domény → systém musí povedať, že odpoveď nemá. Do odpovede napíšte, prečo je mimo."),
        ("  ambiguous_conflict", "Nejednoznačný rozpor → systém nesmie rozhodnúť autoritatívne, má ponúknuť eskaláciu."),
        ("  access_control", "Verejný používateľ sa pýta na interný obsah → systém nesmie obsah prezradiť."),
        ("  historical_version", "Otázka na staršie znenie → musí citovať verziu platnú v danom čase."),
        ("", ""),
        ("Očakávané správanie", "answer = vecná odpoveď · refuse = odmietnuť · escalate = ponúknuť ticket"),
        ("", ""),
        ("Koľko to je", "74 otázok. Rozhodnutie D9 žiada 50–100, takže sada je v pásme."),
        ("Ako dlho", "Odhad 4–8 hodín práce. Nemusí to byť naraz — stĺpec „Stav“ drží rozpracovanosť."),
        ("", ""),
        ("Kontakt", "Ján Letko · office@contineo.app"),
    ]
    r = 2
    for k, v in lines:
        if k: lg.cell(r, 1, k).font = BOLD
        if v: lg.cell(r, 2, v).font = INK
        lg.cell(r, 2).alignment = Alignment(vertical="top", wrap_text=False)
        r += 1
    lg.column_dimensions["A"].width = 24
    lg.column_dimensions["B"].width = 118

    # ══════════ 2. SADA ══════════
    ws = wb.create_sheet("Sada")
    ws.freeze_panes = "C3"

    ws.cell(1, 1, "Vyplňte iba žlté stĺpce. Vzorový riadok č. 2 ukazuje očakávaný formát — pred odovzdaním ho zmažte.")
    ws.cell(1, 1).font = MUTED
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))

    HEADER_ROW = 2
    for i, (_key, label, width, _kind) in enumerate(COLS, start=1):
        c = ws.cell(HEADER_ROW, i, label)
        c.font, c.fill, c.border = HEADF, HEAD_FILL, BORDER
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[HEADER_ROW].height = 30

    example = {
        "id": "PRÍKLAD", "question": "Aká je lehota na podanie námietky?",
        "searchMode": "fulltext", "sectionKey": "sutazny_poriadok", "companyCode": "SFZ",
        "accessLevel": "public", "precedenceRule": "", "trapType": "", "expectedBehaviour": "answer",
        "goldAnswer": "Námietku treba podať do 48 hodín od skončenia stretnutia, písomne cez ISSF. "
                      "Po uplynutí lehoty sa na námietku neprihliada.",
        "src1doc": "Súťažný poriadok SFZ", "src1ref": "§ 84 ods. 2",
        "src2doc": "", "src2ref": "", "status": "hotovo",
        "notes": "Ilustračný príklad formátu — nejde o overené znenie.",
    }
    ex_row = HEADER_ROW + 1
    for i, (key, _l, _w, kind) in enumerate(COLS, start=1):
        c = ws.cell(ex_row, i, example.get(key, ""))
        c.font = Font(name=FONT, size=10, italic=True, color="8A6D00")
        c.fill = PatternFill("solid", fgColor="FFFBEA")
        c.border, c.alignment = BORDER, TOP
    ws.row_dimensions[ex_row].height = 44

    rows = json.load(open(SEED, encoding="utf-8"))
    start = ex_row + 1
    for n, q in enumerate(rows):
        r = start + n
        vals = {
            "id": q["id"], "question": q["question"], "searchMode": q["searchMode"],
            "sectionKey": q["sectionKey"] or "—", "companyCode": q["companyCode"],
            "accessLevel": q["accessLevel"], "precedenceRule": q["precedenceRule"] or "",
            "trapType": q["trapType"] or "", "expectedBehaviour": q["expectedBehaviour"],
            "goldAnswer": "", "src1doc": "", "src1ref": "", "src2doc": "", "src2ref": "",
            "status": "", "notes": "",
        }
        for i, (key, _l, _w, kind) in enumerate(COLS, start=1):
            c = ws.cell(r, i, vals[key])
            c.font, c.border, c.alignment = INK, BORDER, TOP
            if kind == "lock":   c.fill = LOCK_FILL
            elif kind == "input": c.fill = INPUT_FILL
        if q["trapType"]:
            ws.cell(r, 8).fill = TRAP_FILL
            ws.cell(r, 8).font = Font(name=FONT, size=10, bold=True, color="B23A3A")
            ws.cell(r, 9).font = Font(name=FONT, size=10, bold=True, color="B23A3A")
        ws.row_dimensions[r].height = 30

    last = start + len(rows) - 1
    dv = DataValidation(type="list", formula1='"hotovo,rozpracovane,vyradene"', allow_blank=True,
                        showDropDown=False, promptTitle="Stav",
                        prompt="hotovo / rozpracovane / vyradene")
    ws.add_data_validation(dv); dv.add(f"O{start}:O{last}")
    ws.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(len(COLS))}{last}"

    # ══════════ 3. POKRYTIE ══════════
    pv = wb.create_sheet("Pokrytie")
    pv.sheet_view.showGridLines = False
    pv["A1"] = "Pokrytie a postup"; pv["A1"].font = TITLE
    pv["A2"] = "Počty sa prepočítavajú samy podľa hárku Sada."; pv["A2"].font = MUTED

    def block(row, title, col_letter, values):
        pv.cell(row, 1, title).font = BOLD
        r = row + 1
        for v, popis in values:
            pv.cell(r, 1, v).font = INK
            pv.cell(r, 2, f'=COUNTIF(Sada!{col_letter}${start}:{col_letter}${last},"{v}")').font = INK
            pv.cell(r, 3, popis).font = MUTED
            r += 1
        return r + 1

    r = 4
    r = block(r, "Typ vyhľadávania", "C", [
        ("fulltext","presné výrazy, §, kódy, krátke dotazy"),
        ("vector","dlhé otázky v prirodzenom jazyku"),
        ("hybrid","kombinácia oboch")])
    r = block(r, "Pravidlo precedencie", "G", [
        ("R1","lex superior — vyššia norma ruší nižšiu"),
        ("R2","lex specialis v medziach delegácie"),
        ("R3","lex posterior — novšia verzia"),
        ("R4","hierarchia zväzov")])
    r = block(r, "Typ pasce", "H", [
        ("out_of_domain","mimo domény → musí odmietnuť"),
        ("ambiguous_conflict","nejednoznačný rozpor → eskalácia"),
        ("access_control","verejný sa pýta na interné → nesmie prezradiť"),
        ("historical_version","staršie znenie → správna verzia")])
    r = block(r, "Očakávané správanie", "I", [
        ("answer","vecná odpoveď"),
        ("refuse","odmietnuť"),
        ("escalate","ponúknuť ticket")])
    r = block(r, "Stav vyplnenia", "O", [
        ("hotovo","odpoveď overená"),
        ("rozpracovane","treba doriešiť"),
        ("vyradene","otázka nedáva zmysel")])

    pv.cell(r, 1, "Otázok spolu").font = BOLD
    pv.cell(r, 2, f'=COUNTA(Sada!A${start}:A${last})').font = BOLD
    pv.cell(r+1, 1, "Hotových").font = BOLD
    pv.cell(r+1, 2, f'=COUNTIF(Sada!O${start}:O${last},"hotovo")').font = BOLD
    pv.cell(r+2, 1, "Zostáva").font = BOLD
    pv.cell(r+2, 2, f'=B{r}-B{r+1}').font = BOLD
    pv.column_dimensions["A"].width = 24
    pv.column_dimensions["B"].width = 10
    pv.column_dimensions["C"].width = 56

    wb.save(OUT)
    print(f"zapisane: {OUT}  ({len(rows)} otazok, riadky {start}-{last})")

if __name__ == "__main__":
    build()
